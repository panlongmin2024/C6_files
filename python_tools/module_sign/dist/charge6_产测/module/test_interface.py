import base64
import json
import os
import subprocess
import traceback
from datetime import date

from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from module.post_data import *


def handle_response(func):
    def wrapper(*args, **kwargs):
        response = func(*args, **kwargs)
        if response.status_code:
            response_msg = response.json()
            if response.status_code == 200:
                if response_msg['code'] == 200:
                    if 'data' in response_msg:
                        return True, str(response_msg)
            return False, str(response_msg)
        return False, '网络问题'

    return wrapper


def excel_function(func):
    def decorator(*args, **kwargs):
        excel_title, excel_dict = func(*args, **kwargs)
        date01: date = date.today()
        check_result: bool = os.path.exists(path=excel_title)
        print("8")
        if check_result is True:
            wb: Workbook = load_workbook(filename=excel_title)
            sh = wb[str(object=date01)]
            value_data = list(excel_dict.values())
            n_rows = sh.max_row  # 获去最后一行行数
            for i in range(len(value_data)):
                sh.cell(n_rows + 1, i + 1).value = value_data[i]
            # 保存文件
            wb.save(filename=excel_title)
        else:
            # 创建一个Excel workbook 对象
            book = Workbook()
            # 创建时，会自动产生一个sheet，通过active获取
            sh = book.active
            sh.title = str(object=date01)
            sh = book[sh.title]
            first_row = list(excel_dict.keys())
            value_data = list(excel_dict.values())
            for i in range(len(first_row)):
                sh.cell(1, i + 1).value = first_row[i]
                sh.cell(2, i + 1).value = value_data[i]
            # for i in range(len(value_data)):
            #     sh.cell(2, i + 1).value = value_data[i]
            # 保存文件
            book.save(filename=excel_title)

    return decorator


class ThreadInterface(QThread):
    test_end_signal = pyqtSignal(str, str, str)

    def __init__(self):
        super().__init__()
        self.sn: str = '1'
        self.is_mes: bool = False
        self.config_data: dict = {"sn_length": 1}
        self.token: str = ''

    @handle_response
    def pass_station(self):
        url: str = "http://{}/api/mes-product/public/station/center/test/check".format(
            self.config_data['web_url'])

        response = PostTestData().pass_station(
            url=url, config_data=self.config_data, token=self.token, sn=self.sn)
        return response

    @excel_function
    def create_excel(self, result, notes):
        excel_dict: dict = {
            "测试结果": result,
            "备注": notes
        }
        excel_title: str = './excel_files/' + str(date.today()) + '.xlsx'
        return excel_title, excel_dict

    @handle_response
    def send_result(self, test_data):
        machineTestData = {
            "machineCode": self.config_data["machineCode"],
            "titles": {
                "uuid": "100",
                "hashValue": "200",
                "key": "300",
            },
            "results": test_data
        }
        response = PostTestData().upload_data(
            result=machineTestData, post_data=self.config_data,
            sequenceNumber=self.sn, token=self.token
        )
        return response

    def get_sha256_hash(self, file_path):
        cmd = ['openssl', 'dgst', '-sha256', file_path]
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True,
                                universal_newlines=True)
        hash_str = result.stdout.strip().split('\n')[0]
        hash_value = hash_str.split('=')[1].strip()
        return hash_value

    def send_file_for_signing(self, server_url, file_path):
        try:
            # hash_string = self.get_sha256_hash(file_path)
            hash_string = '9267FCFCF1B2FEFB5BBA658E478D455584E42325'
            self.hash_string = hash_string
            print(hash_string)
            data = {
                "hashOfFirmwareImage": hash_string,
                "productModelName": "JBL Test Speaker"
            }
            print(data)
            response = requests.post(server_url, json=data)
            if response.status_code == 200:
                print("request ok\n")
                try:
                    response_data = response.json()
                    if 'signature' in response_data:
                        # print("sign data: ", response_data['signature'])
                        signature = response_data['signature']
                        binary_data = base64.b64decode(signature)
                        return binary_data
                    else:
                        self.test_end_signal.emit("FAIL", "FAIL", "error: no signature key word in response")
                except json.JSONDecodeError:
                    self.test_end_signal.emit("FAIL", "FAIL", "error: not jason data")
            else:
                self.test_end_signal.emit("FAIL", "FAIL", "request failed, status code: " + str(response.status_code))
            return None
        except requests.RequestException as e:
            self.test_end_signal.emit("FAIL", "FAIL", f"An error occurred while sending data for signing: {e}")
            return None

    def sign_file(self, file_path, output_path):
        # if os.path.exists(output_path):
        #     print(f"sign file %s existed" % (output_path))
        #     return
        signature = self.send_file_for_signing(self.server_url, file_path)
        with open(output_path, 'wb') as signature_file:
            signature_file.write(signature)


def thread_decorator(cls):
    class NewClass(QThread):
        test_end_signal: pyqtSignal = pyqtSignal(str, str, str)
        msg_signal: pyqtSignal = pyqtSignal(str)

        # 修改前置操作
        def __init__(self, *args, **kwargs):
            super().__init__()
            self.oc = cls(*args, **kwargs)

        # 修改原始类的方法
        def run(self):
            try:
                if self.oc.sn_length == len(self.oc.sn):  # 如果sn长度相同
                    pass_result, msg = self.oc.pass_station()  # 过站检查
                    if pass_result is True:  # 如果检查通过
                        test_result, test_data, test_msg = self.oc.run()  # QThread run函数
                        if test_result is True:  # 测试通过
                            if self.oc.is_mes is True:  # 如果mes上传为真
                                http_result, msg = self.oc.send_result(test_data)  # 结果上传
                                self.oc.create_excel("PASS", msg)  # 生成excel
                                if http_result is True:  # 如果上传成功
                                    return "PASS", "PASS", msg
                                else:  # 上传失败
                                    return "FAIL", "FAIL", msg
                            else:  # mes上传为假
                                self.oc.create_excel("PASS", test_msg)
                                return "PASS", "PASS", test_msg
                        else:  # 测试不通过
                            self.oc.create_excel("PASS", 'str(e)')
                            return "FAIL", "FAIL", test_msg
                    else:  # 检查不通过
                        return "FAIL", "FAIL", msg
                else:  # 如果sn长度不同
                    return "sn长度不符", "FAIL", ""
            except Exception as e:
                self.oc.create_excel("PASS", str(e))
                return "FAIL", "FAIL", str(e)

    return NewClass
