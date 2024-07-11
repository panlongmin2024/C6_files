import base64
import binascii
import gc
import hashlib
import os
import socket
import sys
import threading
import time
import traceback
import tracemalloc
from binascii import a2b_hex
from datetime import date
import requests
from requests.auth import HTTPBasicAuth
from serial import Serial
from module.serial_method import get_port_list
from ui_py.set_ui import *
from PyQt5 import QtGui, QtCore
from module.get_set_control_value import *
from ui_py.main_window_ui import *
import transitions
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import QApplication, QMainWindow
from transitions import Machine
from module.post_data import PostTestData
from module.logg_module import *
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from module.write_read_json import *


def excel_function(excel_title: str, excel_dict: dict):
    date01: date = date.today()
    check_result: bool = os.path.exists(path=excel_title)
    if check_result is True:
        wb: Workbook = load_workbook(filename=excel_title)
        sh = wb[str(date01)]
        value_data = list(excel_dict.values())
        n_rows = sh.max_row  # 获去最后一行行数
        for i in range(len(value_data)):
            sh.cell(n_rows + 1, i + 1).value = value_data[i]
        # 保存文件
        wb.save(filename=excel_title)
    else:
        # 创建一个Excel workbook 对象
        book = Workbook()
        # 创建时，会自动产生一个sheet，通过active获取
        sh = book.active
        sh.title = str(date01)
        sh = book[sh.title]
        first_row = list(excel_dict.keys())
        value_data = list(excel_dict.values())
        for i in range(len(first_row)):
            sh.cell(1, i + 1).value = first_row[i]
            sh.cell(2, i + 1).value = value_data[i]
        # 保存文件
        book.save(filename=excel_title)


def handle_response(response):
    if response.status_code:
        response_msg = response.json()
        if response.status_code == 200:
            if response_msg['code'] == 200:
                if 'data' in response_msg:
                    return 'PASS', str(response_msg)
        return 'FAIL', str(response_msg)
    return 'FAIL', '网络问题'


def write_hex_in(com: Serial, hex_instr: str) -> None:
    com.write(a2b_hex(hex_instr))


def write_str_in(com: Serial, text: str) -> None:
    com.write(a2b_hex(text.encode().hex()))


class SetFont(QDialog):
    send_value_signal = pyqtSignal(dict)
    save_signal = pyqtSignal(dict)

    def __init__(self, parent=None):
        super(SetFont, self).__init__(parent)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        upload_icon = QIcon('./icon/upload.svg')
        self.setWindowIcon(upload_icon)
        self.setWindowModality(QtCore.Qt.ApplicationModal)
        self.config_data = RWJson.read_json('./configs/config_data.json')
        self.ui.confirm.clicked.connect(self.save_control_value)
        self.ui.resect.clicked.connect(self.clear_all)
        self.setWindowTitle('设置')
        self.ui.baud.setMaximum(5000000)

    def showEvent(self, a0: QtGui.QShowEvent) -> None:
        set_control_value(self, self.config_data)
        self.ui.port.addItems(get_port_list())

    def send_control_value(self) -> None:
        control_values: dict = get_control_value(self, self.config_data)
        self.send_value_signal.emit(control_values)

    def save_control_value(self) -> None:
        control_values: dict = get_control_value(self, self.config_data)
        RWJson.write_json('./configs/config_data.json', control_values)
        self.save_signal.emit(control_values)

    def clear_all(self) -> None:
        control = self.findChildren(QLineEdit)
        for i in control:
            i.clear()


class StateThread(QThread):
    change_text_signal: pyqtSignal = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_continue: bool = False
        self.text_list: list = ['Testing   ',
                                'Testing.  ', 'Testing.. ', 'Testing...']

    def run(self):
        index: int = 0
        while self.is_continue is True:
            self.change_text_signal.emit(self.text_list[index])
            self.msleep(100)
            index = 0 if index >= 3 else index + 1


class Matter(QThread):
    msg_signal = pyqtSignal(str)
    test_end_signal = pyqtSignal(str, str, str)
    restart_thread_signal = pyqtSignal(str, str, str)

    def __init__(self):
        super().__init__()
        self.sn: str = 'A4B65D020020'
        self.config_data: dict = {}
        self.token: str = ''
        self.sn_check_result: bool = False
        self.station_check_result: bool = False
        self.is_mes: bool = True
        self.test_result: str = ''
        self.test_data: list[dict] = []
        self.test_message: str = ''
        self.ip = socket.gethostbyname(socket.gethostname())
        self.check_key_log: str = '未抓到验签log'
        self.com: Serial = Serial()
        # self.ip = os.popen('hostname -I').readline().replace('\\n', '')
        self.machine: Machine = Machine(
            model=self,
            states=['before_get_permit', 'before_test',
                    'after_test', 'after_upload', 'after_create_excel'],
            initial='before_test',
        )
        self.machine.add_transition(trigger='try_get_permit', source='before_get_permit', dest='before_test',
                                    before='get_permit')
        self.machine.add_transition(trigger='test_sth', source='before_test', dest='after_test', before='test_main2',
                                    conditions='test_check')
        self.machine.add_transition(trigger='upload', source='after_test', dest='after_upload', before='upload_data',
                                    conditions='upload_check')
        self.machine.add_transition(trigger='save_data', source=['after_upload', 'after_test'],
                                    dest='after_create_excel',
                                    before='create_excel')

    def run(self) -> None:
        try:
            start_time = time.time()
            self.machine.set_state('before_test')
            self.msleep(int(self.config_data['delay'] * 1000))
            # self.try_get_permit()
            self.test_sth()
            if self.get_state() == 'after_test':
                self.upload()
            self.save_data()
            self.test_end_signal.emit(self.test_result, self.test_result, '')
            end_time = time.time()
            self.msg_signal.emit(str(end_time - start_time))
        except transitions.core.MachineError as e:
            self.test_end_signal.emit('WRONG', 'FAIL', str(e))
        except Exception as e2:
            exc_type, exc_value, exc_tb = sys.exc_info()
            logger.error("Extracted traceback from traceback object:")
            extracted_tb = traceback.extract_tb(exc_tb)
            for frame in extracted_tb:
                filename, lineno, func_name, text = frame
                logger.error(
                    f"File : {filename}, Line : {lineno}, Func.Name : {func_name}, Text : {text}")
            self.test_end_signal.emit('WRONG', 'FAIL', str(e2))

    def enter_uart_test(self) -> None:
        write_hex_in(self.com, '7573657220656e7465725f756172745f')
        self.msleep(100)
        write_hex_in(self.com, '746573740d')

    def get_uuid(self) -> str:
        write_hex_in(self.com, '574c545f524541445f555549440D')
        self.msleep(200)
        raw_uuid: str = str(self.com.read(45))
        uuid2: str = raw_uuid[15:-1]
        self.msg_signal.emit(raw_uuid)
        self.msg_signal.emit(uuid2)
        return uuid2

    def write_key_in(self, signature: str) -> str:
        key_cmd: str = 'WLT_HASH_UUID=' + signature
        key_hex: str = key_cmd.encode().hex()
        self.msg_signal.emit('key_cmd: ' + key_cmd)
        self.msg_signal.emit('write key in')
        self.msg_signal.emit('instruct: ' + str(a2b_hex(key_hex)))
        self.com.write(a2b_hex(key_hex))
        write_result = str(self.com.read(10))
        self.msg_signal.emit(write_result)
        return write_result

    def set_port(self) -> None:
        self.com.baudrate = self.config_data['baud']
        self.com.port = self.config_data['port']
        self.com.stopbits = self.config_data['stop_bit']
        if self.com.isOpen() is False:
            self.com.open()
        self.com.set_buffer_size(2048, 2048)
        self.msg_signal.emit('open port')

    def send_hash_for_signing(self, server_url: str, hash_value: str, uuid):
        self.msg_signal.emit('hash_value:' + hash_value)
        data = [{
            "hashOfHardwareId": hash_value,
            "productModelName": "JBL Charge 6",
          
            #"signatureScheme": "RSASSA-PSS",
            "hardwareId": uuid}]
        self.msg_signal.emit(str(data))
        self.msg_signal.emit("向树莓派发送请求,url: " + server_url)
        response = requests.post(
            url=server_url, json=data,
            auth=HTTPBasicAuth(self.config_data['pi_user_name'], self.config_data['pi_password']),
            verify=self.config_data['cert_path']
        )
        self.msg_signal.emit("response:" + response.text)
        if response.status_code == 200:
            response_data = response.json()
            if 'signature' in response_data[0]:
                signature = response_data[0]['signature']
                return self.change_to_hex(signature), ''
            else:
                return None, "error: no signature uuid_cert word in response"
        else:
            self.msg_signal.emit("5")
            return None, response.text

    def test_main2(self):
        self.set_port()
        self.enter_uart_test()
        self.msleep(300)
        self.com.flushInput()
        uuid = self.get_uuid()
        hash_value: str = self.sha256_hash(uuid)
        signature, self.test_message = self.send_hash_for_signing(self.config_data['pi_url'], hash_value, uuid)
        self.msg_signal.emit("response_msg:"+self.test_message)
        if signature is not None:
            self.com.flushInput()
            write_result = self.write_key_in(signature)
            if 'SUCCESS' in write_result:
                #self.restart_read()
                self.test_message='写入并验签成功'
                self.test_result='PASS'
            else:
                self.test_message = '未获得key值写入成功信息'
                self.test_result = 'FAIL'
        self.com.close()
        self.test_data.append(
            {"uuid": uuid, "hash_value": hash_value, "key": signature, "验签信息": self.check_key_log})

    def restart_read(self):
        t1 = threading.Thread(target=self.keep_read, args=(self.com,))
        #t2 = threading.Thread(target=self.turn_off, args=(self.com,))
        t1.start()
        #t2.start()
        t1.join()
        #t2.join()

    def test_main(self) -> None:
        self.com.baudrate = self.config_data['baud']
        self.com.port = self.config_data['port']
        self.com.stopbits = self.config_data['stop_bit']
        if self.com.isOpen() is False:
            self.com.open()
        self.com.set_buffer_size(2048, 2048)
        self.msg_signal.emit('open port')
        # 'user enter_uart_test'
        self.com.write(a2b_hex('7573657220656e7465725f756172745f'))
        self.msleep(100)
        self.com.write(a2b_hex('746573740d'))
        self.msleep(300)
        self.com.flushInput()
        # 'WLT_READ_UUID'
        self.com.write(a2b_hex('574c545f524541445f555549440D'))
        self.msleep(200)
        raw = self.com.read(45)
        uuid_org = str(raw)
        self.msg_signal.emit(uuid_org)
        uuid2 = uuid_org[15:-1]
        self.msg_signal.emit(uuid2)
        self.msg_signal.emit("uuid_string:" + uuid2)
        hash_value: str = self.sha256_hash(uuid2)
        signature, msg = self.send_hash_for_signing(self.config_data['pi_url'], hash_value)
        self.com.flushInput()
        # random_string = str(uuid.uuid4()).replace('-', '')[:512]
        key_cmd: str = 'WLT_HASH_UUID=' + signature
        # key_cmd=
        key_hex: str = key_cmd.encode().hex()
        self.msg_signal.emit('key_cmd: ' + key_cmd)
        self.msg_signal.emit('instruct: ' + key_hex)
        self.msg_signal.emit('write key in')
        self.com.write(a2b_hex(key_hex))
        self.msg_signal.emit(str(self.com.read(10)))
        # self.restart_thread_signal.emit(uuid2, hash_value, signature)
        t1 = threading.Thread(target=self.keep_read, args=(self.com,))
        t2 = threading.Thread(target=self.turn_off, args=(self.com,))
        t1.start()
        t2.start()
        t1.join()
        t2.join()
        self.test_data.append(
            {"uuid": uuid2, "hash_value": hash_value, "key": signature, "验签信息": self.check_key_log})

    def change_to_hex(self, base64_str: str) -> str:
        # 解码base64字符串
        decoded_data = base64.b64decode(base64_str)
        # 将解码后的数据转换为16进制字符串
        hex_str = binascii.hexlify(decoded_data).decode('utf-8')
        self.msg_signal.emit('hex key: ' + hex_str)
        self.msg_signal.emit('hex key len: ' + str(len(hex_str)))
        return hex_str

    def sha256_hash(self, s: str) -> str:
        return hashlib.sha256(s.encode('utf-8')).hexdigest()

    def keep_read(self, com: Serial) -> None:
        self.msg_signal.emit('Keep read com')
        # 开始时间
        start_time = time.time()
        while True:
            if com.in_waiting > 0:
                msg: bytes = com.readline()
                print(msg)
                if b'uuid verify ok' in msg:
                    self.msg_signal.emit(str(msg))
                    self.test_message = '验签成功'
                    self.check_key_log = str(msg)[2:-7]
                    self.test_result = 'PASS'
                    break
                elif b'uuid verify fail' in msg:
                    self.msg_signal.emit(str(msg))
                    self.test_message = '验签失败'
                    self.check_key_log = str(msg)
                    self.test_result = 'FAIL'
                    break
                # 检查是否已经过去了5秒
                if time.time() - start_time >= 5:
                    self.test_result = 'FAIL'
                    self.test_message = '未抓到验签log'
                    break
        self.msg_signal.emit(self.check_key_log)
        self.msg_signal.emit(self.test_message)
        self.msg_signal.emit(self.test_result)

    def turn_off(self, com: Serial) -> None:
        com.flushInput()
        self.msg_signal.emit('Restart')
        write_str_in(self.com, 'TL_DUT_REBOOT')
        # com.write(data=a2b_hex('TL_DUT_REBOOT'.encode().hex()))

    def get_permit(self):
        self.msg_signal.emit("是否启用MES： " + str(self.is_mes))
        if self.is_mes is True:
            response = PostTestData().post_permit(post_data=self.config_data)
            response_json = response.json()
            self.msg_signal.emit(str(response_json))
            if response.status_code == 200:
                self.token = response_json['access_token']
                return True
            else:
                return False
        return True

    def test_check(self):
        if len(self.sn) == self.config_data['sn_length']:  # 如果sn长度相同
            if self.is_mes is True:
                self.msg_signal.emit('发送过站请求: ')
                pass_result, msg = self.pass_station()  # 过站检查
                self.msg_signal.emit(msg)
                if pass_result == 'FAIL':
                    self.test_end_signal.emit('过站失败', 'FAIL', '过站失败')
                    return False
                return True
            return True
        else:
            self.test_end_signal.emit('sn长度不符', 'FAIL', 'sn长度不符')
            return False

    def pass_station(self):
        url: str = "http://{}/api/mes-product/public/station/center/test/check".format(
            self.config_data['web_url'])
        response = PostTestData().pass_station(
            url=url, config_data=self.config_data, token=self.token, sn=self.sn)
        return handle_response(response)

    def upload_data(self):
        machine_test_data = {
            "machineCode": self.config_data["machineCode"],
            "titles": {
                "uuid": "100",
                "hashValue": "200",
                "key": "300",
            },
            "results": self.test_data
        }
        response = PostTestData().upload_data(
            result=machine_test_data, post_data=self.config_data,
            sequenceNumber=self.sn, token=self.token
        )
        # self.msg_signal.emit(str(response.json()))
        self.test_result, self.test_message = handle_response(response)

    def upload_check(self):
        self.is_mes = True if self.test_result == "PASS" and self.is_mes is True else False
        return self.is_mes

    def create_excel(self):
        excel_dict: dict = {}
        for i in self.test_data:
            for j in i:
                excel_dict[j] = i[j]
        excel_dict['结果'] = self.test_result
        excel_dict['备注'] = self.test_message
        excel_title: str = './excel_files/' + str(date.today()) + '.xlsx'
        excel_function(excel_title, excel_dict)
        # self.test_result=str(self.test_result)

    def get_state(self):
        return self.machine.get_model_state(self).value


class Restart(QThread):
    msg_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.com: Serial = Serial()
        self.type = ''

    def run(self) -> None:
        try:

            self.com.flushInput()
            self.msg_signal.emit('restart')
            self.com.write(a2b_hex('TL_DUT_REBOOT'.encode().hex()))
        except Exception as e:
            self.msg_signal.emit(str(e))


class KeepRead(QThread):
    test_end_signal = pyqtSignal(str, str, str)
    msg_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.config_data: dict = {}
        self.token: str = ''
        self.sn: str = ''
        self.is_mes = ''
        self.com: Serial = Serial()
        self.uuid: str = ''
        self.key: str = ''
        self.hash_value: str = ''
        self.check_key_log: str = ''
        self.test_result: str = ''
        self.test_message: str = ''
        self.test_data: list = []

    def run(self) -> None:
        try:
            self.keep_read()
            self.upload_check()
            if self.is_mes is True:
                self.upload_data()
            self.create_excel()
            self.test_end_signal.emit(
                self.test_result, self.test_result, self.test_message)
        except Exception as e:
            self.msg_signal.emit(str(e))

    def keep_read(self):
        self.com.port = self.config_data['port']
        self.com.baudrate = self.config_data['baud']
        self.com.stopbits = self.config_data['stop_bit']
        self.msg_signal.emit('Keep read')
        # 开始时间
        start_time = time.time()
        while True:
            if self.com.in_waiting > 0:
                msg = self.com.readline()
                if b'uuid verify ok' in msg:
                    self.test_result = 'PASS'
                    result_data = str(msg)[2:-7]
                    self.check_key_log = 'uuid verify ok'
                    self.test_message = '验签成功'
                    break
                elif b'uuid verify fail' in msg:
                    self.test_result = 'FAIL'
                    self.test_message = '验签失败'
                    self.check_key_log = 'uuid verify fail'
                    break
                # 检查是否已经过去了5秒
                if time.time() - start_time >= 5:
                    self.test_result = 'FAIL'
                    self.test_message = '验签失败'
                    self.check_key_log = '未抓到验签log'
                    break
        self.msg_signal.emit(self.check_key_log)
        self.msg_signal.emit(self.test_result)
        self.test_data = [
            {"uuid": self.uuid, "hash_value": self.hash_value, "key": self.key, "验签信息": self.check_key_log}]

    def upload_data(self):
        machine_test_data = {
            "machineCode": self.config_data["machineCode"],
            "titles": {
                "uuid": "100",
                "hashValue": "200",
                "key": "300",
            },
            "results": self.test_data
        }
        response = PostTestData().upload_data(
            result=machine_test_data, post_data=self.config_data,
            sequenceNumber=self.sn, token=self.token
        )
        # self.msg_signal.emit(str(response.json()))
        self.test_result, self.test_message = handle_response(response)

    def upload_check(self):
        self.is_mes = True if self.test_result == "PASS" and self.is_mes is True else False
        self.msg_signal.emit("是否启用MES： " + str(self.is_mes))
        return self.is_mes

    def create_excel(self):
        excel_dict: dict = {}
        for i in self.test_data:
            for j in i:
                excel_dict[j] = i[j]
        excel_dict['结果'] = self.test_result
        excel_dict['备注'] = self.test_message
        excel_title: str = './excel_files/' + str(date.today()) + '.xlsx'
        excel_function(excel_title, excel_dict)
        # self.test_result=str(self.test_result)


class UiInterface(QMainWindow):
    def __init__(self):
        super().__init__()
        gc.enable()
        tracemalloc.start()
        self.ui: Ui_MainWindow = Ui_MainWindow()
        self.ui.setupUi(self)
        self.token: str = ''
        self.config_data: dict = RWJson.read_json('./configs/config_data.json')
        self.setting_font: SetFont = SetFont()
        self.ui.setting.clicked.connect(self.setting_font.show)
        self.setting_font.save_signal.connect(
            lambda config_data: self.refresh_upload_config(config_data))
        self.get_permit()

    def refresh_upload_config(self, config_data: dict):
        self.config_data = config_data
        self.get_permit()
        self.setting_font.close()

    def get_permit(self) -> None:
        try:
            if self.ui.is_mes is True:
                response = PostTestData().post_permit(post_data=self.config_data)
                response_json = response.json()
                if response.status_code == 200:
                    self.token = response_json['access_token']
                    self.ui.log.append_log('token获取成功')
                    self.setting_font.close()
                else:
                    self.ui.log.append_log(response_json)
                    QMessageBox.critical(self.setting_font, '错误', 'token获取失败')
        except Exception as e:
            print(e)

    def start_test(self):
        pass

    def test_end_handel(self, state: str, pass_or_fail: str, message: str):
        pass

    def quit_test(self):
        pass

    def get_test_state(self):
        pass

    def set_font(self):
        font: QFont = QFont()
        font.setWeight(600)
        font.setPointSize(18)
        font.setFamily(u"Microsoft New Tai Lue")
        self.ui.test_state.setMinimumSize(QSize(250, 42))
        self.ui.test_state.setFont(font)
        font.setPointSize(25)
        font.setBold(True)
        self.ui.company_name.setFont(font)

    def clear_trash(self):
        gc.collect()  # 手动回收垃圾
        current_mem, peak_mem = tracemalloc.get_traced_memory()
        self.ui.log.append_log(f"released, Current memory usage is {current_mem / 10 ** 6}MB " +
                               f"Peak was {peak_mem / 10 ** 6}MB")
        self.ui.log.append_log(
            '======================================================================')

    def set_icon(self):
        window_icon = QIcon('./icon/上位机.ico')
        self.setWindowIcon(window_icon)
        setting_icon = QIcon('./icon/setting.svg')
        self.ui.setting.setIcon(setting_icon)
        stop_icon = QIcon('./icon/stop.svg')
        self.ui.stop.setIcon(stop_icon)


class UseMatter(UiInterface):
    def __init__(self):
        super().__init__()
        self.is_mes_dict: dict = RWJson.read_json('./configs/is_mes.json')
        self.state_thread: StateThread = StateThread()
        self.state_thread.change_text_signal.connect(
            lambda text: self.ui.test_state.setText(text))
        self.test_thread: Matter = Matter()
        self.test_thread.msg_signal.connect(self.ui.log.append_log)
        self.test_thread.test_end_signal.connect(self.test_end_handel)
        # self.test_thread.restart_thread_signal.connect(self.restart_signal_handle)
        self.ui.sn.returnPressed.connect(self.start_test)
        self.ui.stop.clicked.connect(self.quit_test)
        # set_control_value(self, self.is_mes_dict)
        set_control_value(self.ui.num_widget, self.is_mes_dict)
        self.ui.num_widget.test_num1 = int(self.is_mes_dict['test_num'])
        self.ui.num_widget.pass_num1 = int(self.is_mes_dict['pass_num'])
        self.config_data = RWJson.read_json('./configs/config_data.json')
        self.setWindowTitle(self.config_data['title_line'])
        self.ui.company_name.setText(self.config_data['company_line'])
        self.ui.company_name.setAlignment(Qt.AlignCenter)
        self.set_font()
        self.set_icon()
        self.ui.test_state.setAlignment(Qt.AlignCenter)
        # self.ui.num_widget.test_num.setText('5')
        # self.restart_thread:Restart = Restart()
        # self.restart_thread.msg_signal.connect(self.ui.log.append_log)
        # self.keep_read_thread :KeepRead= KeepRead()
        # self.keep_read_thread.msg_signal.connect(self.ui.log.append_log)
        # self.keep_read_thread.test_end_signal.connect(self.test_end_handel)
        # self.com = ''

    # def restart_signal_handle(self, uuid2, hash_value, signature):
    #     self.keep_read_thread.config_data = self.config_data
    #     self.keep_read_thread.is_mes = self.test_thread.is_mes
    #     self.keep_read_thread.uuid = uuid2
    #     self.keep_read_thread.hash_value = hash_value
    #     self.keep_read_thread.key = signature
    #     self.keep_read_thread.sn = self.test_thread.sn
    #     # self.restart_thread.com.port=self.config_data['port']
    #     # self.restart_thread.com.baudrate=self.config_data['baud']
    #     # self.restart_thread.com.stopbits=self.config_data['stop_bit']
    #     self.restart_thread.com = self.com
    #     self.keep_read_thread.com = self.com
    #     self.keep_read_thread.start()
    #     self.restart_thread.start()

    def closeEvent(self, a0: QtGui.QCloseEvent) -> None:
        self.quit_state_thread()
        control_values: dict[str] = {
            'test_num': self.ui.num_widget.test_num.text(),
            'pass_num': self.ui.num_widget.pass_num.text(),
            'passing_rate': self.ui.num_widget.passing_rate.text()
        }
        RWJson.write_json('./configs/is_mes.json', control_values)

    def get_test_state(self):
        return self.test_thread.isRunning()

    def start_test(self):
        try:
            self.ui.sn.setEnabled(False)
            self.ui.test_state.setStyleSheet(u"background-color:#FF9900;\n"
                                             "border:1px solid black;")
            self.ui.log.clear_log()
            self.state_thread.is_continue = True
            self.test_thread.sn = self.ui.sn.text()
            self.test_thread.config_data = self.config_data
            self.test_thread.token = self.token
            self.test_thread.is_mes = self.ui.is_mes.isChecked()
            self.state_thread.start()
            self.test_thread.start()
        except Exception as e:
            self.ui.log.append_log(str(e))

    def quit_state_thread(self):
        if self.state_thread.isRunning() is True:
            self.state_thread.is_continue = False
            self.state_thread.terminate()

    def quit_test_thread(self):
        if self.state_thread.isRunning() is True:
            self.test_thread.quit()
            self.test_thread.wait()
            self.test_end_handel('Stop testing', 'FAIL', '中断测试')

    def quit_test(self):
        # self.test_thread.terminate()
        # self.quit_test()
        self.quit_state_thread()
        self.test_thread.quit()

    def test_end_handel(self, state: str, pass_or_fail: str, message: str):
        self.quit_state_thread()
        self.ui.test_state.setText(state)
        self.ui.log.append_log(message)
        self.ui.log.create_log(self.ui.sn.text())
        if pass_or_fail == "PASS":
            self.ui.test_state.setStyleSheet(
                "background-color:green;border:1px solid black;")
        else:
            self.ui.test_state.setStyleSheet(
                "background-color:#ff3300;border:1px solid black;")
        if message != 'sn长度不符' and message != '过站失败':
            self.ui.num_widget.handle_result(pass_or_fail)
        self.ui.sn.clear()
        self.ui.sn.setEnabled(True)
        self.ui.sn.setFocus()
        self.clear_trash()


app = QApplication([])
use_matter = UseMatter()
# use_matter.start()
use_matter.show()
app.exec_()
